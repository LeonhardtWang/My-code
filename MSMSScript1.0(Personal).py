import requests
from bs4 import BeautifulSoup
import re
import pandas as pd
import os
import pymongo
import openpyxl
import pickle
import datetime
import time

class Work_Order_Spider:
    '''
    一个爬虫类，用于爬取工单信息
    '''
    def __init__(self):
        '''
        初始化
        '''
        self.session = requests.Session()
        self.login_url = 'http://58.220.2.26:8888/logincheck.php' # 登录的url
        self.login_header = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;\
            q=0.9,image/webp,image/apng,*/*;q=0.8',
            'Accept-Encoding': 'gzip, deflate',
            'Accept-Language': 'zh-CN,zh;q=0.9',
            'Host': '58.220.2.26:8888',
            'Origin': 'http://58.220.2.26:8888',
            'Proxy-Connection': 'keep-alive',
            'Referer': 'http://58.220.2.26:8888/?t=PC',
            'User-Agent': 'Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 \
            (KHTML, like Gecko) Chrome/65.0.3325.181 Mobile Safari/537.36'
        }
        self.username = 'admin' # 账户
        self.password = '' # 密码
        # 下面page_url为提取总页数，总工单数以及RUN_ID和FLOW_ID信息的url
        self.page_url = 'http://58.220.2.26:8888/general/approve_center/query/data/getdata.php?\
        action=list_allwork&flow_id=all&flow_query_type=ALL&flow_status=1&work_level=all&search_begin_dept=all' # flow_status=1表示已结束工单
        self.info_header = {
            'Accept': 'application/json, text/javascript, */*; q=0.01',
            'Accept-Encoding': 'gzip, deflate',
            'Accept-Language': 'zh-CN,zh;q=0.9',
            'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
            'Host': '58.220.2.26:8888',
            'Origin': 'http://58.220.2.26:8888',
            'Proxy-Connection': 'keep-alive',
            'Referer': 'http://58.220.2.26:8888/general/approve_center/query/',
            'User-Agent': 'Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 \
            (KHTML, like Gecko) Chrome/65.0.3325.181 Mobile Safari/537.36',
            'X-Requested-With': 'XMLHttpRequest'
        }
        self.info_base_url = 'http://58.220.2.26:8888/general/approve_center/list/print.php?FLOW_VIEW=12' # 提取特定工单信息的url
        self.permis_lis = ['33', '49', '50', '63', '61'] # 注意：输入的类型号！！！！！！！！！

    def login(self):
        '''
        登录
        '''
        post_data = {
            'UNAME': self.username,
            'PASSWORD': self.password
        }
        self.session.post(self.login_url, data=post_data, headers=self.login_header)   

    def run_id_flow_id(self):
        '''
        获取所有的RUN_ID和FLOW_ID,保存进属性id_list中
        参数：
        view_record：是否只返回总的工单数目，默认False
        '''
        res = self.session.post(self.page_url, headers=self.info_header)           
        page_tol = res.json()['total'] # 提取总页数
        
        self.id_list = []
        for page in range(1, page_tol+1):
            post_data = {
                'page': page
            }
            res = self.session.post(self.page_url, data=post_data, headers=self.info_header)
            for row in res.json()['rows']:
                regex = re.compile(r'("","(\d+?)","","(\d+?)")') # 提取RUN_ID和FLOW_ID的正则表达式
                re_res = re.findall(regex, row['cell'][2])[0]
                run_id, flow_id = re_res[-2], re_res[-1]
                self.id_list.append((run_id, flow_id))

    def deal_excel(self, content, res, flow_id):
        '''
        判断excel类型
        '''
        if flow_id in self.permis_lis:
            with open('temp.xlsx', 'wb') as f:
                f.write(res.content)
            df = pd.read_excel('temp.xlsx', sheet_name='Sheet1')
            os.remove('temp.xlsx')
            # 以下判断是否有交换机->端口键
            col_lis = list(df.columns)
            col_new = []
            cursor = 0
            sing_index = []
            contate_index = []
            while cursor <= len(col_lis) - 1:
                if col_lis[cursor] == '对端交换机':
                    start = cursor
                    cursor += 4
                elif col_lis[cursor] == '端口':
                    end = cursor
                    cursor += 4
                elif col_lis[cursor] == '新对端交换机':
                    start = cursor
                    cursor += 4
                elif col_lis[cursor] == '新端口':
                    end = cursor
                    cursor += 4
                else:
                    col_new.append(col_lis[cursor])
                    sing_index.append(cursor)
                    cursor += 1
            if '对端交换机' in col_lis and '端口' in col_lis:
                col_new.extend(['端口1', '端口2', '端口3', '端口4'])
                contate_index = [(start+i, end+i) for i in range(4)]
            if '新对端交换机' in col_lis and '新端口' in col_lis:
                col_new.extend(['新端口1', '新端口2', '新端口3', '新端口4'])
                contate_index = [(start+i, end+i) for i in range(4)]
        
            for _, row in df.iterrows():
                col_cursor = 0
                item_excel = {}
                for i in sing_index:
                    if pd.isnull(row[i]):
                        col_cursor += 1
                        continue
                    item_excel[col_new[col_cursor]] = str(row[i])
                    col_cursor += 1
                if contate_index:
                    for src, dst in contate_index:
                        if pd.isnull(row[src]) or pd.isnull(row[dst]):
                            col_cursor += 1
                            continue
                        item_excel[col_new[col_cursor]] = str(row[src]) + str('->') + str(row[dst])
                        col_cursor += 1
                content_copy = content.copy() 
                content_copy.update(item_excel)
                self.content_lis.append(content_copy)

            if not self.content_lis:
                self.content_lis.append(content)

            

    def get_work_order_info(self, run_id, flow_id):
        '''
        获取特定RUN_ID和FLOW_ID工单的信息，存入content
        参数：
        run_id:RUN_ID
        flow_id:FLOW_ID
        返回：
        content:类型List
        '''
        info_url = self.info_base_url + '&RUN_ID=' + run_id + '&FLOW_ID=' + flow_id
        res = self.session.get(info_url, headers=self.info_header)
        content = {}
        soup_form = BeautifulSoup(res.text, 'lxml').select('tbody td') # 解析器为lxml
        td_key_value = ((soup_form[i], soup_form[i+1]) for i in range(0, len(soup_form)-1, 2))
        link_excel = re.findall(re.compile(r'<a href="(/inc/attach.php?[^/\.]+?\.xlsx)">下载</a>', re.S), res.text)
        self.content_lis = [] # 仅在有excel上传文件时，它才有超过一个的条目

        content['run_id'] = run_id; content['flow_id'] = flow_id
        for key, value in td_key_value:
            key_text = key.text.replace(u'\xa0', u' ').replace(u'&nbsp', u' ').replace(u'<br />', u' ').strip()
            value_text = value.text.replace(u'\xa0', u' ').replace(u'&nbsp', u' ').replace(u'<br />', u' ').strip()
            content[key_text] = value_text
        

        if link_excel:
            excel_url = 'http://58.220.2.26:8888' + link_excel[0].replace('&amp;', '&')
            res = self.session.get(excel_url, headers=self.info_header)

            self.deal_excel(content, res, flow_id)

        if not self.content_lis:
            self.content_lis.append(content)

    def to_mongodb(self):
        self.login()
        self.run_id_flow_id()

        client = pymongo.MongoClient(host='localhost', port=27017)      
        #client = pymongo.MongoClient(host='58.220.2.26', port=27017)
        #client = pymongo.MongoClient(host='58.220.2.26', port=27017, username='root', password='9787534348432wxy')
        db = client['test']
        #collections = db['workorder']

        work_order_id = []
        carry_excel_id = []
        for con in self.id_list:
            if con[1] in self.permis_lis: # 如果添了特定处理方式的类型号，这里添if
                carry_excel_id.append((con[0], con[1]))
            else:
                work_order_id.append((con[0], con[1]))

        # 标上唯一ID
        if os.path.exists('ID_start.pkl'):
            with open('ID_start.pkl', 'rb') as f:
                ID = pickle.load(f)
        else:
            # 查看数据库中是否有数据，有数据则备份
            for db_na in client.list_database_names():
                if db_na in ['admin', 'local']:continue
                time_now = str(datetime.datetime.now()).split('.')[0].split(' ')
                time_now = time_now[0] + '#'+ time_now[1]
                time_now = time_now.replace(':', '-')
                for clc in client[db_na].list_collection_names():
                    rlt_lis = list(client[db_na][clc].find())
                    if rlt_lis:
                        db_now = client[time_now]
                        db_now[clc].insert_many(rlt_lis) # 备份
                        client[db_na].drop_collection(clc)
            ID = 1


        # 标上唯一SNID
        if os.path.exists('SNID_start.pkl'):
            with open('SNID_start.pkl', 'rb') as f:
                SNID = pickle.load(f)
        else:
            SNID = 1

        
        # 提取carry_excel表中SN和SNID键值对
        collect_carry_excel = db['carry_excel']
        results = list(collect_carry_excel.find())
        SNID_carry_excel = {}
        if results:
            for res in results:
                SNID_carry_excel[res['SN']] = res['SNID']


        compare_data = CompareData()
        # 初处理库中work_order表增删
        work_order_id_add = compare_data.work_order_add(work_order_id)
        work_order_id_del = compare_data.work_order_del(work_order_id)
        if work_order_id_add:
            for run_id, flow_id in work_order_id_add:
                collections = db['work_order']
                self.get_work_order_info(run_id, flow_id)
                for item in self.content_lis:
                    item['ID'] = str(ID) # 新增ID键

                    item_copy = item.copy()
                    collections.insert_one(item_copy)
                    ID += 1
        if work_order_id_del:
            for run_id, flow_id in work_order_id_del:
                collections = db['work_order']
                collections.delete_many({'run_id':run_id, 'flow_id':flow_id})
        
        # 初处理库中carry_excel表增删
        carry_excel_id_add = compare_data.carry_excel_add(carry_excel_id)
        carry_excel_id_del = compare_data.carry_excel_del(carry_excel_id)
        if carry_excel_id_add:
            for run_id, flow_id in carry_excel_id_add:
                collections = db['carry_excel']
                self.get_work_order_info(run_id, flow_id)
                for item in self.content_lis:
                    item['ID'] = str(ID) # 新增ID键

                    # 新增SNID键
                    if item['flow_id'] == '61':
                        pass
                    elif item['SN'] in SNID_carry_excel:
                        item['SNID'] = SNID_carry_excel[item['SN']]
                    else:
                        item['SNID'] = str(SNID)
                        SNID_carry_excel[item['SN']] = item['SNID']
                        SNID += 1

                    item_copy = item.copy()
                    collections.insert_one(item_copy)
                    ID += 1
        if carry_excel_id_del:
            for run_id, flow_id in carry_excel_id_del:
                collections = db['carry_excel']
                collections.delete_many({'run_id':run_id, 'flow_id':flow_id})

        with open('ID_start.pkl', 'wb') as f:
            pickle.dump(ID, f)

        with open('SNID_start.pkl', 'wb') as f:
            pickle.dump(SNID, f)

class CompareData:
    def __init__(self):
        client = pymongo.MongoClient(host='localhost', port=27017)      
        #client = pymongo.MongoClient(host='58.220.2.26', port=27017)
        #client = pymongo.MongoClient(host='58.220.2.26', port=27017, username='root', password='9787534348432wxy')
        db_test = client['test']
        db_user = client['设备资产']
        
        # 初处理库中work_order表的id
        self.test_db_work_order_lis = []
        result = db_test['work_order'].find()
        for res in result:
            self.test_db_work_order_lis.append((res['run_id'], res['flow_id']))

        # 初处理库中carry_excel表的id
        self.test_db_carry_excel_lis = []
        result = db_test['carry_excel'].find()
        for res in result:
            self.test_db_carry_excel_lis.append((res['run_id'], res['flow_id']))
        
        # 处理用户库中'设备资产变更表'的id
        self.user_db_alter_lis = []
        result = db_user['设备资产变更表'].find()
        for res in result:
            self.user_db_alter_lis.append((res['工单号'], res['操作类型']))
    
    # 初处理库中word_order表的增删id
    def work_order_add(self, id_now_lis):
        return list(set(id_now_lis) - set(self.test_db_work_order_lis))
    def work_order_del(self, id_now_lis):
        return list(set(self.test_db_work_order_lis) - set(id_now_lis))

    # 初处理库中carry_excel表的id
    def carry_excel_add(self, id_now_lis):
        return list(set(id_now_lis) - set(self.test_db_carry_excel_lis))
    def carry_excel_del(self, id_now_lis):
        return list(set(self.test_db_carry_excel_lis) - set(id_now_lis))
    
    # 处理用户库中'设备资产变更表'的id
    def alter_asset_add(self, id_now_lis):
        return list(set(id_now_lis) - set(self.user_db_alter_lis))
    def alter_asset_del(self, id_now_lis):
        return list(set(self.user_db_alter_lis) - set(id_now_lis))
    
class WanderExcel:
    def __init__(self):
        self.alter_col = ['ID',  'SNID', '工单号', '发单时间', '操作类型', '客户名称', '机房', \
            '机架位置', '设备种类','设备简称','SN', '设备型号', 'IP地址', '端口1', \
                '端口2', '端口3', '端口4', '设备状态']

        #self.ol_order_col = ['工单号', '客户名称', '机架位置', '设备种类', '设备简称', 'SN', \
        #    '设备型号', 'IP地址', '端口1', '端口2', '端口3', '端口4'] # 基于设备资产变更表生成设备资产线上表

        client = pymongo.MongoClient(host='localhost', port=27017)      
        #client = pymongo.MongoClient(host='58.220.2.26', port=27017)
        #client = pymongo.MongoClient(host='58.220.2.26', port=27017, username='root', password='9787534348432wxy')
        self.db = self.client['test']
        self.user_db = self.client['设备资产']
        self.collect = self.db['work_order']
        self.collect_xlsx = self.db['carry_excel']
        self.flow_id = {'49':'上架', '50':'下架', '63':'搬迁', '61': '修改'} # '33'情况不在里面

    def alter_order(self, to_excel=False): # 设备资产变更表
        results = self.collect_xlsx.find()
        res_lis = []
        for res in results:
            item_dict = {}
            for col in self.alter_col:
                if col == '操作类型':
                    if col in res:
                        item_dict[col] = res[col]
                    else:
                        item_dict[col] = self.flow_id[res['flow_id']]
                elif col == '工单号':
                    if 'run_id' in res:
                        item_dict[col] = res['run_id']
                elif col == 'IP地址':
                    if 'IP地址' in res:
                        item_dict[col] = res['IP地址']
                    elif '新IP地址' in res:
                        item_dict[col] = res['新IP地址']
                elif col == '端口1':
                    if '端口1' in res:
                        item_dict[col] = res['端口1']
                    elif '新端口1' in res:
                        item_dict[col] = res['新端口1']
                elif col == '端口2':
                    if '端口2' in res:
                        item_dict[col] = res['端口2']
                    elif '新端口2' in res:
                        item_dict[col] = res['新端口2']
                elif col == '端口3':
                    if '端口3' in res:
                        item_dict[col] = res['端口3']
                    elif '新端口3' in res:
                        item_dict[col] = res['新端口3']
                elif col == '端口4':
                    if '端口4' in res:
                        item_dict[col] = res['端口4']
                    elif '新端口4' in res:
                        item_dict[col] = res['新端口4']
                elif col == '机架位置':
                    if '新机架位置' in res:
                        item_dict[col] = res['新机架位置'] 
                    elif '机架位置' in res:
                        item_dict[col] = res['机架位置']
                else:
                    if col in res:
                        item_dict[col] = res[col]
            res_lis.append(item_dict)
        alter_asset_id = [(res['工单号'], res['操作类型']) for res in res_lis]
        compare_data = CompareData()
        alter_asset_id_add = compare_data.alter_asset_add(alter_asset_id)
        alter_asset_id_del = compare_data.alter_asset_del(alter_asset_id)

        collections = self.user_db['设备资产变更表']
        item_lis = []
        # 增加条目
        for res in res_lis:
            if (res['工单号'], res['操作类型']) in alter_asset_id_add:
                item_lis.append(res)
        if item_lis:
            collections.insert_many(item_lis)

        # 删除条目
        if alter_asset_id_del:
            for run_id, flow_id in alter_asset_id_del:
                collections.delete_many({'工单号':run_id, '操作类型':flow_id})
        # 生成excel
        #if to_excel:
        #    df = pd.DataFrame(res_lis)
        #    df.to_excel('设备资产变更表.xlsx', index=False)

    def ol_order(self): # 设备资产线上表
        index_dict = {}
        item_lis = []
        collections = self.user_db['设备资产线上表']
        collections_alter_asset = self.user_db['设备资产变更表']
        results = collections_alter_asset.find()
        results = sorted(results, key=lambda s:int(s['工单号']))
        for i, res in enumerate(results):
            if res['操作类型'] == '修改':
                if res['设备状态'] in ['线上', '离线']:
                    index_dict[res['SNID']] = i
            elif res['操作类型'] in ['上架', '搬迁']:
                if res['SNID'] not in index_dict:
                    if res['操作类型'] == '搬迁':
                        print("工单号：%s，SNID：%s，SN：%s出错！\n此设备不在线上!\n"\
                            % (res['工单号'], res['SNID'], res['SN']))
                        continue
                    index_dict[res['SNID']] = i
                else:
                    if res['操作类型'] == '上架':
                        print("工单号：%s, SNID：%s, SN：%s出错！\n此设备已上架！\n"\
                            % (res['工单号'], res['SNID'], res['SN']))
                        continue
                    index_dict[res['SNID']] = i
            elif res['操作类型'] == '下架':
                try:
                    del index_dict[res['SNID']]
                except:
                    print("工单号：%s，SNID：%s, SN：%s出错！\n此设备不在线上！\n" \
                        % (res['工单号'], res['SNID'], res['SN']))
        
        for i in index_dict.values():
            del results[i]['发单时间']
            del results[i]['操作类型']
            item_lis.append(results[i])
        collections.delete_many({})         
        if item_lis:
            collections.insert_many(sorted(item_lis, \
                key=lambda s:int(s['工单号']), reverse=True))

    def store_asset(self):# 设备资产库存表
        index_dict = {}
        item_lis = []
        collections = self.user_db['设备资产库存表']
        collections_alter_asset = self.user_db['设备资产变更表']
        results = collections_alter_asset.find()
        results = sorted(results, key=lambda s:int(s['工单号']))

        for i, res in enumerate(results):
            if res['操作类型'] == '修改': # 判断修改
                if res['设备状态'] == '库存':
                    index_dict[res['SNID']] = i
            if res['操作类型'] in ['签收', '下架']:
                if res['操作类型'] == '签收':
                    if res['操作类型'] in index_dict:
                        print("工单号：%s，SNID：%s, SN：%s出错！\n此设备已在库存！\n"\
                            % (res['工单号'], res['SNID'], res['SN']))
                    else:
                        index_dict[res['SNID']] = i
                else:
                    index_dict[res['SNID']] = i
            elif res['操作类型'] in ['发送', '上架']:
                try:
                    del index_dict[res['SNID']]
                except:
                    print("工单号：%s，SNID：%s，SN：%s出错！\n此设备不在库存！\n"\
                        % (res['工单号'], res['SNID'], res['SN']))
        
        for i in index_dict.values():
            for key in ['发单时间', '操作类型', '机架位置', 'IP地址', \
                '端口1', '端口2', '端口3', '端口4']:
                try:
                    del results[i][key]
                except:
                    continue
            item_lis.append(results[i])
        collections.delete_many({})
        if item_lis:
            collections.insert_many(sorted(item_lis, \
                key=lambda s:int(s['工单号']), reverse=True))

    def net_port_use(self): # 网络设备端口使用表
        item_lis = []
        col_name = []
        collections = self.user_db['网络设备端口使用表']
        collections_ol_order = self.user_db['设备资产线上表']

        results = collections_ol_order.find()
        results = list(results)
        if results:
            port_key_val = []
            for res in results:
                if res['设备种类'] == '网络设备':
                    col_name.append(res['设备简称'])
                if '端口1' in res:
                    port_key_val.append(res['端口1'])
                if '端口2' in res:
                    port_key_val.append(res['端口2'])
                if '端口3' in res:
                    port_key_val.append(res['端口3'])
                if '端口4' in res:
                    port_key_val.append(res['端口4'])
            
            k = 0
            while port_key_val:
                item = {}
                if k <= len(col_name) -1:
                    item['网络设备'] = col_name[k]
                    k += 1
                for col in col_name:
                    if not port_key_val:break
                    for i, k_v in enumerate(port_key_val):
                        if k_v.split('->')[0] == col:
                            item[col] = k_v.split('->')[1]
                            del port_key_val[i]
                            break
                item_lis.append(item)
            while k <= len(col_name) - 1:
                item = {}
                item['网络设备'] = col_name[k]
                item_lis.append(item)
                k += 1
            collections.delete_many({})
            collections.insert_many(item_lis)

    def IP_use(self): # IP地址使用表
        collections = self.user_db['IP地址使用表']
        collections_ol_order = self.user_db['设备资产线上表']
        results = collections_ol_order.find()
        collections.delete_many({})
        for res in results:
            if 'IP地址' not in res:continue
            for ip in res['IP地址'].split('/'):
                if not ip:continue
                item = {}
                item['已使用IP'] = ip
                collections.insert_one(item)


def convertToTitle(n):
    """
    输入数字返回excel列号
    :type n: int
    :rtype: str
    """
    res = ''
    while n:
        n, mod = divmod(n, 26)
        if mod == 0:
            res = 'Z' + res
            n -= 1
        else:
            res = chr(64 + mod) + res
    return res

class ServeForExcel:
    def __init__(self, excel_name, exc_col, sheet_name, db_name, collect_name, complete=False):
        self.excel_name = excel_name
        df = pd.read_excel(excel_name, sheet_name=sheet_name)
        self.shape = df.shape
        self.exc_col = exc_col
        #self.exc_col = {'SN':'A', '设备种类':'B', '设备简称':'C', '设备型号':'D'}
        self.wb = openpyxl.load_workbook(excel_name)
        self.ws = self.wb[sheet_name]

        client = pymongo.MongoClient(host='localhost', port=27017)      
        #client = pymongo.MongoClient(host='58.220.2.26', port=27017)
        #client = pymongo.MongoClient(host='58.220.2.26', port=27017, username='root', password='9787534348432wxy')
        self.user_db = self.client[db_name]
        self.collect = self.user_db[collect_name]
        self.complete = complete

    def clear(self):
        row_start = 2
        if self.complete:
            row_start = 1    
        for i in range(1, self.shape[1]+1):
            for j in range(2, self.shape[0]+2):
                self.ws[convertToTitle(i)+str(j)] = ''
    
    def write(self):
        results = self.collect.find()
        results = list(results)
        if self.complete:
            for i in range(len(self.exc_col)):
                self.ws[convertToTitle(i+1)+'1'] = list(self.exc_col.keys())[i]
        row = 2
        for res in results:
            for key in self.exc_col.keys():
                if key in res:
                    self.ws[self.exc_col[key]+str(row)] = res[key]
            row += 1
    
    def save(self):
        self.wb.save(self.excel_name)
    
    def run(self):
        self.clear()
        self.write()
        self.save()
    
class Log:
    '''
    日志输出
    '''
    pass
if __name__ == "__main__":
    spider = Work_Order_Spider()
    spider.to_mongodb()
    wander_excel = WanderExcel()
    wander_excel.alter_order()
    wander_excel.ol_order()
    wander_excel.store_asset()
    wander_excel.net_port_use()
    wander_excel.IP_use()


    # 处理设备发送填写模板
    excel_name = '设备发送填写模板.xlsx'
    df = pd.read_excel(excel_name, sheet_name='Sheet2')
    exc_col = {}
    for i, c in enumerate(df.columns):
        exc_col[c] = convertToTitle(i+1)
    sheet_name = 'Sheet2'
    db_name = '设备资产'
    collect_name = '设备资产库存表'
    serve_for_excel = ServeForExcel(excel_name, exc_col, sheet_name,\
        db_name, collect_name)
    serve_for_excel.run()


    # 处理设备下架填写模板
    excel_name = '设备下架填写模板.xlsx'
    df = pd.read_excel(excel_name, sheet_name='Sheet2')
    exc_col = {}
    for i, c in enumerate(df.columns):
        exc_col[c] = convertToTitle(i+1)
    sheet_name = 'Sheet2'
    db_name = '设备资产'
    collect_name = '设备资产线上表'
    serve_for_excel = ServeForExcel(excel_name, exc_col, sheet_name,\
        db_name, collect_name)
    serve_for_excel.run()


    # 处理设备上架填写模板
    excel_name = '设备上架填写模板.xlsx'
    df = pd.read_excel(excel_name, sheet_name='Sheet2')
    exc_col = {}
    for i, c in enumerate(df.columns):
        exc_col[c] = convertToTitle(i+1)
    sheet_name = 'Sheet2'
    db_name = '设备资产'
    collect_name = '设备资产库存表'
    serve_for_excel = ServeForExcel(excel_name, exc_col, sheet_name,\
        db_name, collect_name)
    serve_for_excel.run()

    df = pd.read_excel(excel_name, sheet_name='Sheet3')
    exc_col = {}
    for i, c in enumerate(df.columns):
        exc_col[c] = convertToTitle(i+1)
    sheet_name = 'Sheet3'
    collect_name = 'IP地址使用表'
    serve_for_excel = ServeForExcel(excel_name, exc_col, sheet_name,\
        db_name, collect_name)
    serve_for_excel.run()

    client = pymongo.MongoClient(host='localhost', port=27017)      
    #client = pymongo.MongoClient(host='58.220.2.26', port=27017)
    #client = pymongo.MongoClient(host='58.220.2.26', port=27017, username='root', password='9787534348432wxy')
    user_db = client['设备资产']
    collection = user_db['网络设备端口使用表']
    exc_col = {}
    res_max = []
    results = collection.find()
    results = list(results)
    for i, res in enumerate(results):
        if len(res)  > len(res_max):
            res_max = res
    del res_max['_id']
    for i, c in enumerate(res_max.keys()):
        exc_col[c] = convertToTitle(i+1)

    sheet_name = 'Sheet4'
    collect_name = '网络设备端口使用表'
    serve_for_excel = ServeForExcel(excel_name, exc_col, sheet_name, \
        db_name, collect_name,complete=True)
    serve_for_excel.run()


    # 处理设备搬迁填写模板
    excel_name = '设备搬迁填写模板.xlsx'
    df = pd.read_excel(excel_name, sheet_name='Sheet2')
    exc_col = {}
    for i, c in enumerate(df.columns):
        exc_col[c] = convertToTitle(i+1)
    sheet_name = 'Sheet2'
    db_name = '设备资产'
    collect_name = '设备资产线上表'
    serve_for_excel = ServeForExcel(excel_name, exc_col, sheet_name,\
        db_name, collect_name)
    serve_for_excel.run()

    df = pd.read_excel(excel_name, sheet_name='Sheet3')
    exc_col = {}
    for i, c in enumerate(df.columns):
        exc_col[c] = convertToTitle(i+1)
    sheet_name = 'Sheet3'
    collect_name = 'IP地址使用表'
    serve_for_excel = ServeForExcel(excel_name, exc_col, sheet_name,\
        db_name, collect_name)
    serve_for_excel.run()

    client = pymongo.MongoClient(host='localhost', port=27017)      
    #client = pymongo.MongoClient(host='58.220.2.26', port=27017)
    #client = pymongo.MongoClient(host='58.220.2.26', port=27017, username='root', password='9787534348432wxy')
    user_db = client['设备资产']
    collection = user_db['网络设备端口使用表']
    exc_col = {}
    res_max = []
    results = collection.find()
    results = list(results)
    for i, res in enumerate(results):
        if len(res)  > len(res_max):
            res_max = res
    del res_max['_id']
    for i, c in enumerate(res_max.keys()):
        exc_col[c] = convertToTitle(i+1)

    sheet_name = 'Sheet4'
    collect_name = '网络设备端口使用表'
    serve_for_excel = ServeForExcel(excel_name, exc_col, sheet_name, \
        db_name, collect_name,complete=True)
    serve_for_excel.run() 


    # 处理设备信息更改表
    excel_name = '设备信息更改模板.xlsx'
    df = pd.read_excel(excel_name, sheet_name='Sheet3')
    exc_col = {}
    for i, c in enumerate(df.columns):
        exc_col[c] = convertToTitle(i+1)
    sheet_name = 'Sheet3'
    db_name = '设备资产'
    collect_name = '设备资产线上表'
    serve_for_excel = ServeForExcel(excel_name, exc_col, sheet_name,\
        db_name, collect_name)
    serve_for_excel.run() 

    df = pd.read_excel(excel_name, sheet_name='Sheet4')
    exc_col = {}
    for i, c in enumerate(df.columns):
        exc_col[c] = convertToTitle(i+1)
    sheet_name = 'Sheet4'
    db_name = '设备资产'
    collect_name = '设备资产库存表'
    serve_for_excel = ServeForExcel(excel_name, exc_col, sheet_name,\
        db_name, collect_name)
    serve_for_excel.run() 

    df = pd.read_excel(excel_name, sheet_name='Sheet5')
    exc_col = {}
    for i, c in enumerate(df.columns):
        exc_col[c] = convertToTitle(i+1)
    sheet_name = 'Sheet5'
    db_name = '设备资产'
    collect_name = 'IP地址使用表'
    serve_for_excel = ServeForExcel(excel_name, exc_col, sheet_name,\
        db_name, collect_name)
    serve_for_excel.run() 


    df = pd.read_excel(excel_name, sheet_name='Sheet6')
    exc_col = {}
    for i, c in enumerate(df.columns):
        exc_col[c] = convertToTitle(i+1)
    sheet_name = 'Sheet6'
    db_name = '设备资产'
    collect_name = '网络设备端口使用表'
    serve_for_excel = ServeForExcel(excel_name, exc_col, sheet_name,\
        db_name, collect_name)
    serve_for_excel.run() 





    

